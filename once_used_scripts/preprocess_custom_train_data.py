import openpyxl
import os
import re

# létrehoztam az érzelmek integer alapú kódolásához egy szótárat, ahol a string érzelmeket számokhoz rendeltem
emotion_dictionary = {
    "Neutral": 0,
    "0": 0,
    "Fear": 1,
    "fear": 1,
    "Sadness": 2,
    "sadness": 2,
    "Anger": 3,
    "anger": 3,
    "Disgust": 4,
    "disgust": 4,
    "Success": 5,
    "success": 5,
    "Joy": 6,
    "joy": 6,
    "Trust": 7,
    "trust": 7
}


# névelem "lecsupaszítása" egy függvénnyel
def format_name_element(param):
    # először töröljük a zárójeles kifejezést, pld. (ORG)
    regex = re.compile(" [(].*[)]")
    # majd levágjuk a szóközöket a kifejezés elejéről és végéről
    return regex.sub("", param).strip()


# ez a függvény az n-edik előfordulást cseréli a mondatban
# sentence: a mondat, old_value: kicserélendő érték, new_value: új érték, occurence: hanyaik előfordulást kell cserélni
def replace_nth_occurrence(sentence, old_value, new_value, occurrence):
    cnt = 0
    regex = re.escape(old_value)
    for match in re.finditer(regex, sentence):
        if cnt == occurrence:
            corrected_new_value = new_value if match.group() == old_value else new_value + match.group()[-1]
            return sentence[:match.start()] + corrected_new_value + sentence[match.end():]
        cnt += 1
    # ha az old_value nem található a mondatban, üres stringet ad vissza
    return ""


# az input és output fileok elérési útvonalai, alkalmazás esetén átírandó
file_path = "../resources/absa_training set_english_v1.xlsx"
output_path = "../resources/ENG_HunEmPoli_8_ner_valid.txt"

# töltsük be az excelt a programba
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

print(f"Excel sikeresen betöltve. Utolsó sor értéke: {sheet.max_row}")

# ha létezik az output file, akkor töröljük
if os.path.exists(output_path):
    os.remove(output_path)

# létrehozzuk az output file-t
with open(output_path, "w", encoding="utf-8") as file:
    # for ciklussal bejárjuk az excel sorait, kettőtől kezdődik, mert az első sorban a fejléc van,
    # a sentence változóba belerakjul az excel adott sorából a mondatot, ha a mondat cella üres, a ciklus továbblép
    for row in sheet.iter_rows(min_row=2, values_only=True):

        sentence = row[2]
        if sentence is None:
            continue
        # egy másik szótárat használunk a mondaton belüli egyfomra névelemek sorszámához
        name_element_count = {}

        # sornak eddig megnéztük 1 celláját, ami a mondat volt, de most kell a többi cella, hogy az érzelem+névelem párosokat feltudjuk dolgozni
        # E oszlop az első névelem az excelben, ezért indulunk a 4. oszloptól (0-tól kezdődik az indexelés)
        # a ChatGPT rosszul határozta meg a sor hosszát, ezért out of index hibára futott, ezért a 64-es érték
        for j in range(4, 64, 2):

            # ide kerül a j-edik oszlop értéke (névelem)
            name_element_cell = row[j]
            # ide kerül a j+1-edik oszlop értéke (a névelemhez tartozó érzelem)
            emotion_cell = row[j + 1]
            # kiszedjük a dictionaryből az érzelem kódját, ami az outputhoz kell
            emotion_id = emotion_dictionary.get(str(emotion_cell), -1)

            # akkor dolgozunk tovább a sorral, ha van névelem, hozzá tartozó érzelem, és az érzelem szerepel a szótárunkban, különben nem csinálunk semmit az adott mondat-névelem párral
            if name_element_cell != "" and emotion_cell != "" and emotion_id >= 0:
                # névelem formázzuk
                name_element_trimmed = format_name_element(name_element_cell)
                # ha a névelem megtalálható a mondatban, cseréljük
                if name_element_trimmed in sentence:
                    # beállítjuk, hogy ez hanyadik azonos névelem, erre az egy mondatban előforduló azonos névelemek miatt van szükség
                    if name_element_trimmed in name_element_count:
                        name_element_count[name_element_trimmed] += 1
                    else:
                        name_element_count[name_element_trimmed] = 0

                    # cseréljük a mondatban az n-edik névelemet $T$-re
                    sentence_to_print = replace_nth_occurrence(sentence, name_element_trimmed, "$T$",
                                                               name_element_count[name_element_trimmed])
                # ha nem, akkor a mondatot megjelöljük a #### + hibaüzenettel
                else:
                    sentence_to_print = "#### HIBA - Nem találtam meg a névelemet:\n" + sentence

                # szükséges adatok kiírása fájlba
                file.write(sentence_to_print + "\n")
                file.write(name_element_trimmed + "\n")
                file.write(str(emotion_id) + "\n")

                # adatok kiírása konzolra
                # print(sentence_to_print )
                # print(name_element_trimmed)
                # print(str(emotion_id))

print("Feladat végrehajtása sikeres.")